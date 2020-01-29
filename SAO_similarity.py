#coding=utf-8
#Populating the interactive namespace from numpy and matplotlib
# import warnings
# warnings.filterwarnings(action='ignore', category=UserWarning, module='gensim')
#from gensim.models import Word2Vec
'''
import math
from math import sqrt
import numpy
from sklearn.cluster import KMeans
'''

import seaborn as sns
import numpy as np
import random
import matplotlib as mpl
import matplotlib.pyplot as plt
from scipy import stats
import matplotlib as mpl
import matplotlib.pyplot as plt

# style set
sns.set_palette('deep', desat=.6)
sns.set_context(rc={'figure.figsize': (8, 5)})
np.random.seed(1425)
# figsize

errorpos=["a","s","r"]
import json
from nltk.corpus import wordnet as wn
from nltk.corpus import wordnet_ic
import io
from scipy.stats.stats import pearsonr
from openpyxl import load_workbook
from openpyxl import Workbook
#from nltk.corpus import stopwords
brown_ic = wordnet_ic.ic('ic-brown.dat')
semcor_ic=wordnet_ic.ic('ic-semcor.dat')

#filter_words = set(stopwords.words('english'))
filter_words =set([u'just', u'over', u'through',  u'had', u'should', u'to', u'only', u'won', u'under', u'ours', u'them', u'his', u'very', u'they', u'during', u'now', u'him', u'nor', u'd', u'this', u'she', u'each', u'further', u'where', u'there', u'been', u'whom', u'too', u'wouldn', u'themselves', u'until', u'more', u'himself', u'that', u"didn't", u'but', u"that'll", u'with', u'than', u'those', u'he', u'me', u'myself', u'ma', u'these', u'up', u'will', u'while', u'ain', u'can', u'theirs', u'my', u'and', u've', u'then', u'is', u'am', u'it', u'doesn', u'an', u'as', u'itself', u'at', u'in', u'any', u'if', u'again', u'no', u'when', u'same', u'how', u'other', u'which', u'you', u'after', u'most', u'such', u'why', u'a', u'off', u'i', u'm', u'yours', u"you'll", u'so', u'y', u'the', u'once'])

'''
word similareity based on wordnet
'''
def get_wordnet_sim(word1, word2):
    synsets1 = wn.synsets(word1)
    synsets2 = wn.synsets(word2)
    maxx = 0.0
    score=0.0
    scoretemp=[0.0 for all in range(5)]

    for synset1 in synsets1:
        for synset2 in synsets2:
            scoretemp[1] = synset1.wup_similarity(synset2)
            if synset1._pos == synset2._pos:
                tempmax = max(synset1.lch_similarity(synset1), synset2.lch_similarity(synset2))
                tems = synset1.lch_similarity(synset2)
                if tempmax > 0 and tems >= 0:
                    scoretemp[0] = tems / tempmax
            if synset1._pos == synset2._pos and synset1._pos not in errorpos and synset2._pos not in errorpos:
                tempmax3 = max(synset1.res_similarity(synset1,brown_ic), synset2.res_similarity(synset2,brown_ic))
                scoretemp[2] = synset1.res_similarity(synset2,brown_ic)/tempmax3
                tempmax4 = max(synset1.jcn_similarity(synset1,brown_ic), synset2.jcn_similarity(synset2, brown_ic))
                scoretemp[3] = synset1.jcn_similarity(synset2,brown_ic)/tempmax4
                scoretemp[4] = synset1.lin_similarity(synset2,semcor_ic)


            #average score


            tempscore=0.0
            countnonzero=0
            for i in range(len(scoretemp)):
                if scoretemp[i]>0 and scoretemp[i]<=1:
                    tempscore+=scoretemp[i]
                    countnonzero+=1
            if countnonzero==0:
                return 0.0
            else:
                score = tempscore/countnonzero


            #score = sum(scoretemp)/5

            #score=max(scoretemp)

            if score is not None:
                if score > maxx:
                    maxx = score
    return maxx
'''
compute the f_identify
'''
def f_identify(x, score_level):
    scorelen=len(score_level)
    if x>1 or x<0:
        print x
        raise Exception("Sorry, similarity score is not in [0,1]!")
    if x==1:
        return 1
    for i in range(scorelen-1):
        if x>=score_level[i] and x<score_level[i+1]:
            return (score_level[i]+score_level[i+1])/2
            #return score_level[i+1]

"""
term similarity (the acknowledged method)
"""

def get_term_sim(term1, term2,yuzhi=0.6):
    term1=term1.encode()
    term2=term2.encode()
    term1 = str.lower(term1).replace('\t',' ').replace('-',' ').replace(',','')
    term2 = str.lower(term2).replace('\t', ' ').replace('-',' ').replace(',','')
    term2.strip()
    term1.strip()

    term1=[w for w in term1.split(" ") if not w in filter_words]
    term2=[w for w in term2.split(" ") if not w in filter_words]

    words1 = []
    for word in term1:
        if word is "" or len(wn.synsets(word))==0:continue
        else:
            words1.append(word)

    words2 = []
    for word in term2:
        if word is "" or len(wn.synsets(word))==0:continue
        else:
            words2.append(word)
    match = 0.0

    max_i=len(words1)
    max_j=len(words2)
    if max_i<=0 or max_j<=0:
        direct=[x for x in term1 if x in term2]
        return 2*(float(len(direct)))/(len(term1)+len(term2))

    flag_i=-1
    flag_j=-1

    max_step_num=min(max_i,max_j)

    max_sim=[]
    for k in range(max_step_num):
        max_temp=-1

        for i in range(k,max_i):
            for j in range(k,max_j):
                sim= get_wordnet_sim(words1[i],words2[j])
                if sim > 1 or sim < 0:
                    print sim
                    raise Exception("Sorry, similarity score is not in [0,1]!")
                if sim >max_temp:
                    flag_i=i
                    flag_j=j
                    max_temp=sim

        temp=words1[flag_i]
        words1[flag_i]=words1[k]
        words1[k]=temp
        temp = words2[flag_j]
        words2[flag_j] = words2[k]
        words2[k] = temp
        max_sim.append(max_temp)
        if max_temp >=yuzhi:
            match += 1
    return 2*match/(max_i+max_j)

# term similarity our method

def my_get_term_sim(term1, term2, score_level):
    term1 = term1.encode()
    term2 = term2.encode()
    term1 = str.lower(term1).replace('\t', ' ').replace('-', ' ').replace(',', '')
    term2 = str.lower(term2).replace('\t', ' ').replace('-', ' ').replace(',', '')
    term2.strip()
    term1.strip()

    term1 = [w for w in term1.split(" ") if not w in filter_words]
    term2 = [w for w in term2.split(" ") if not w in filter_words]

    words1 = []
    for word in term1:
        if word is "" or len(wn.synsets(word)) == 0:
            continue
        else:
            words1.append(word)

    words2 = []
    for word in term2:
        if word is "" or len(wn.synsets(word)) == 0:
            continue
        else:
            words2.append(word)
    match = 0.0

    max_i = len(words1)
    max_j = len(words2)
    if max_i <= 0 or max_j <= 0:
        direct = [x for x in term1 if x in term2]
        return 2 * (float(len(direct))) / (len(term1) + len(term2))

    flag_i = -1
    flag_j = -1



    max_step_num = min(max_i, max_j)

    max_sim = []
    for k in range(max_step_num):
        max_temp = -1

        for i in range(k, max_i):
            for j in range(k, max_j):
                sim = get_wordnet_sim(words1[i], words2[j])
                if sim > 1 or sim < 0:
                    print sim
                    raise Exception("Sorry, similarity score is not in [0,1]!")
                if sim > max_temp:
                    flag_i = i
                    flag_j = j
                    max_temp = sim

        temp = words1[flag_i]
        words1[flag_i] = words1[k]
        words1[k] = temp
        temp = words2[flag_j]
        words2[flag_j] = words2[k]
        words2[k] = temp


        max_sim.append(max_temp)

        match += f_identify(max_temp, score_level)

    return 2 * match / (max_i + max_j)

"""
SAO similarity (the acknowledged method)
"""
def get_SAO_sim(SAO1, SAO2, yuzhi_1=0.6, alpha=0.6):

    [s_1, a_1, o_1]=SAO1
    [s_2, a_2, o_2] = SAO2
    sim_ss=get_term_sim(s_1,s_2,yuzhi_1)
    sim_so=get_term_sim(s_1,o_2,yuzhi_1)
    sim_os=get_term_sim(o_1, s_2,yuzhi_1)
    sim_oo=get_term_sim(o_1, o_2,yuzhi_1)
    temp_1=sim_ss+sim_oo
    temp_2=sim_so+sim_os
#    print get_term_sim(a_1, a_2)
    sim_sao12=alpha*((max(temp_1,temp_2))/ 2.0)+(1-alpha)*(get_term_sim(a_1, a_2,yuzhi_1))
#    print sim_sao12
    return sim_sao12
"""
SAO similarity (our method)
"""
def my_get_SAO_sim(SAO1, SAO2, score_level, alpha=0.7):

    [s_1, a_1, o_1]=SAO1
    [s_2, a_2, o_2] = SAO2
    sim_ss=my_get_term_sim(s_1,s_2,score_level)
    sim_so=my_get_term_sim(s_1,o_2,score_level)
    sim_os=my_get_term_sim(o_1, s_2,score_level)
    sim_oo=my_get_term_sim(o_1, o_2,score_level)
    temp_1=sim_ss+sim_oo
    temp_2=sim_so+sim_os
#    print get_term_sim(a_1, a_2)
    sim_sao12=alpha*((max(temp_1,temp_2))/ 2.0)+(1-alpha)*(my_get_term_sim(a_1, a_2,score_level))
#    print sim_sao12
    return sim_sao12

if __name__ == '__main__':


    patenttxt = io.open("sorted_patent.txt", "r", encoding='utf-8')
    patentset = json.loads(patenttxt.read())
    patenttxt.close()

    saosets=[]
    for patent in patentset:
        saosets=saosets+patent[3]
    print len(saosets)

    wb = load_workbook("./similarity score results.xlsx")
    sheet = wb.get_sheet_by_name("human")
    humanjudge=[]
    for i in range(1,len(saosets)):
        humanjudge.append(int(sheet.cell(row=i+1, column=4).value))
    #print humanjudge, len(humanjudge)

    alpha=0.7
    yuzhi_2=0.5

    iii=-1
    our_accurate=[]
    our_precision=[]
    our_recall=[]
    our_fmeasure=[]
    # our method
    wb.create_sheet('ourmethod')
    ws1 = wb['ourmethod']
    wb.save("./similarity score results.xlsx")
    jj=1
    for ii in [3,4,5,10,20,30,40,50,100]:

        ws1.cell(row=1,column=jj,value=ii)


        TP=0
        FN=0
        FP=0
        TN=0

        score_level = [(1.0 / ii) * l for l in range(0, ii+1, 1)] # 分数段，用来确定容错函数
        iii=iii+1

        for i in range(1,len(saosets)):
            aa=my_get_SAO_sim(saosets[0], saosets[i], score_level, alpha)
            ws1.cell(row=i+1, column=jj, value=aa)
            if aa>= yuzhi_2:
                sign=1
            else:
                sign = 0

            if sign==1 and humanjudge[i-1]==1:
                TP+=1
            if sign==0 and humanjudge[i-1]==1:
                FN+=1
            if sign==1 and humanjudge[i-1]==0:
                FP+=1
            if sign==0 and humanjudge[i-1]==0:
                TN+=1


        # compute performance
        our_accurate.append(float(TP+TN)/(len(humanjudge)))
        our_precision.append(float(TP)/(float(TP+FP)))
        our_recall.append(float(TP) / (float(TP + FN)))
        our_fmeasure.append((2*our_precision[iii]*our_recall[iii])/(our_precision[iii]+our_recall[iii]))
        print ii, our_accurate[iii], our_precision[iii], our_recall[iii],our_fmeasure[iii]
        jj += 1
    wb.save("./similarity score results.xlsx")


    iii=-1
    accurate=[]
    precision=[]
    recall=[]
    fmeasure=[]
    # the acknowledged method
    wb.create_sheet('TheAcknowledgedmethod')
    ws2 = wb['TheAcknowledgedmethod']
    jj=1
    for yuzhi_1 in [0.1,0.2,0.3,0.4,0.5,0.6,0.7,0.8,0.9]:
        ws2.cell(row=1,column=jj,value=yuzhi_1)


        iii=iii+1

        TP=0
        FN=0
        FP=0
        TN=0

        for i in range(1,len(saosets)):
            aa=get_SAO_sim(saosets[0], saosets[i], yuzhi_1, alpha)
            ws2.cell(row=i + 1, column=jj, value=aa)
            if aa>= yuzhi_2:
                sign=1
            else:
                sign = 0

            if sign==1 and humanjudge[i-1]==1:
                TP+=1
            if sign==0 and humanjudge[i-1]==1:
                FN+=1
            if sign==1 and humanjudge[i-1]==0:
                FP+=1
            if sign==0 and humanjudge[i-1]==0:
                TN+=1


        # compute the performance
        accurate.append(float(TP + TN) / (len(humanjudge)))
        precision.append(float(TP) / (float(TP + FP)))
        recall.append(float(TP) / (float(TP + FN)))
        fmeasure.append((2 * precision[iii] * recall[iii]) / (precision[iii] + recall[iii]))
        print yuzhi_1, accurate[iii], precision[iii], recall[iii], fmeasure[iii]
        jj += 1
    wb.save("./similarity score results.xlsx")













